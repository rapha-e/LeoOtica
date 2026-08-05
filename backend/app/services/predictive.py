import io
from datetime import datetime, timedelta
from typing import List, Dict, Any
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
import pandas as pd
from backend.app.models.lens import LensInventoryGrade
from backend.app.models.movement import StockMovement

from sqlalchemy import text

async def calculate_predictive_alerts(
    db: AsyncSession,
    lead_time_days: int = 7,
    safety_days: int = 5,
    coverage_days: int = 15
) -> List[Dict[str, Any]]:
    """
    Analisa os consumos e estoque de Lentes Acabadas e Blocos Semiacabados.
    """
    # Consulta a view SQL consolidada para Lentes
    query = text("SELECT * FROM mv_lens_consumption_velocity;")
    res = await db.execute(query)
    rows = res.all()
    
    alerts = []
    
    for row in rows:
        daily_rate = float(row[12])
        safety_stock = daily_rate * safety_days
        reorder_point = (daily_rate * lead_time_days) + safety_stock
        
        current_stock = row[8]
        if current_stock == 0:
            status = "RUPTURA"
        elif current_stock <= reorder_point:
            status = "ALERTA"
        else:
            status = "NORMAL"
            
        suggested = 0
        if daily_rate > 0:
            target_stock = (daily_rate * coverage_days) + safety_stock
            suggested = max(0, int(target_stock - current_stock))
            
        alerts.append({
            "id": str(row[0]),
            "item_type": "LENTE",
            "brand": row[1],
            "material": row[2],
            "refractive_index": float(row[3]),
            "treatment": row[4],
            "diameter": row[5],
            "spherical": float(row[6]),
            "cylindrical": float(row[7]),
            "barcode": row[10] or "N/A",
            "quantity_available": current_stock,
            "location_tag": row[9] or "N/A",
            "total_out_30_days": int(row[11]),
            "daily_consumption_rate": round(daily_rate, 4),
            "reorder_point": round(reorder_point, 2),
            "status": status,
            "suggested_purchase": suggested
        })

    # Consulta a grade de Blocos Semiacabados
    from backend.app.models.block import BlockGridItem
    from sqlalchemy import and_
    block_stmt = select(BlockGridItem).options(selectinload(BlockGridItem.block_model))
    block_items = (await db.execute(block_stmt)).scalars().all()
    
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    for bitem in block_items:
        if not bitem.block_model or not bitem.block_model.is_active:
            continue
        current_stock = bitem.quantity_available
        mov_stmt = select(func.coalesce(func.sum(StockMovement.quantity), 0)).where(
            and_(
                StockMovement.lens_inventory_id == bitem.id,
                StockMovement.movement_type == 'OUT',
                StockMovement.movement_date >= thirty_days_ago
            )
        )
        consumed_30 = (await db.execute(mov_stmt)).scalar() or 0
        daily_rate = float(consumed_30) / 30.0
        safety_stock = daily_rate * safety_days
        reorder_point = (daily_rate * lead_time_days) + safety_stock
        
        if current_stock == 0:
            status = "RUPTURA"
        elif current_stock <= reorder_point:
            status = "ALERTA"
        else:
            status = "NORMAL"
            
        suggested = 0
        if daily_rate > 0:
            target_stock = (daily_rate * coverage_days) + safety_stock
            suggested = max(0, int(target_stock - current_stock))
            
        alerts.append({
            "id": str(bitem.id),
            "item_type": "BLOCO",
            "brand": bitem.block_model.brand,
            "refractive_index": float(getattr(bitem.block_model, 'refractive_index', 1.56) or 1.56),
            "cost_price": float(getattr(bitem.block_model, 'cost_price', 35.00) or 35.00),
            "sale_price": float(getattr(bitem.block_model, 'sale_price', 95.00) or 95.00),
            "treatment": bitem.block_model.name,
            "diameter": "N/A",
            "spherical": float(bitem.base_curve),
            "cylindrical": float(bitem.addition),
            "barcode": bitem.barcode or "N/A",
            "quantity_available": current_stock,
            "location_tag": bitem.location_tag or "N/A",
            "total_out_30_days": int(consumed_30),
            "daily_consumption_rate": round(daily_rate, 4),
            "reorder_point": round(reorder_point, 2),
            "status": status,
            "suggested_purchase": suggested
        })
        
    return alerts


def generate_purchase_plan_excel(alerts_data: List[Dict[str, Any]]) -> bytes:
    """
    Gera uma planilha Excel contendo as sugestões de compra para itens
    que estão em RUPTURA ou ALERTA, formatada no padrão visual elegante do PDF.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Filtra apenas itens com recomendação de compra > 0
    filtered_data = [item for item in alerts_data if item["suggested_purchase"] > 0]
    
    # Se não houver itens com sugestão, gera com todos para não ir vazio
    if not filtered_data:
        filtered_data = alerts_data
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Sugestões de Reposição"
    
    # Garante que as linhas de grade estejam visíveis
    ws.views.sheetView[0].showGridLines = True
    
    # Estilos
    font_family = "Segoe UI"
    
    # Preenchimentos
    fill_header = PatternFill(start_color="2A6BB8", end_color="2A6BB8", fill_type="solid")
    fill_info_box = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
    fill_tag = PatternFill(start_color="EBF2FA", end_color="EBF2FA", fill_type="solid")
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    fill_note = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    
    # Fontes
    font_logo = Font(name=font_family, size=18, bold=True, color="1F4E78")
    font_logo_sub = Font(name=font_family, size=9, bold=False, color="595959")
    font_tag = Font(name=font_family, size=9, bold=True, color="2A6BB8")
    font_sec_title = Font(name=font_family, size=11, bold=True, color="2A6BB8")
    font_sec_val = Font(name=font_family, size=10, bold=False, color="1F2937")
    font_sec_val_bold = Font(name=font_family, size=10, bold=True, color="1F2937")
    font_tbl_header = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    font_tbl_cell = Font(name=font_family, size=10, bold=False, color="1F2937")
    font_tbl_cell_bold = Font(name=font_family, size=10, bold=True, color="1F2937")
    font_tbl_cell_blue = Font(name=font_family, size=10, bold=True, color="2A6BB8")
    font_tbl_cell_red = Font(name=font_family, size=10, bold=True, color="DC2626")
    font_tbl_cell_orange = Font(name=font_family, size=10, bold=True, color="D97706")
    font_note_text = Font(name=font_family, size=9, bold=False, color="78350F")
    
    # Alinhamentos
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    align_note = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Bordas
    border_thin = Side(style="thin", color="E5E7EB")
    border_info = Side(style="thin", color="CBD5E1")
    
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # 1. LOGO E TITULO DO MÓDULO (Linhas 2 a 3)
    ws.merge_cells("B2:D2")
    ws["B2"] = "Nova Lab"
    ws["B2"].font = font_logo
    ws["B2"].alignment = align_left
    
    ws.merge_cells("B3:D3")
    ws["B3"] = "MÓDULO DE AUTOMAÇÃO DE ESTOQUE"
    ws["B3"].font = font_logo_sub
    ws["B3"].alignment = align_left
    
    # Tag de Pedido de Reposição Automática
    ws.merge_cells("G2:I3")
    ws["G2"] = "PEDIDO DE REPOSIÇÃO AUTOMÁTICA"
    ws["G2"].font = font_tag
    ws["G2"].fill = fill_tag
    ws["G2"].alignment = align_center
    
    # Desenha borda fina na tag G2:I3
    for r in range(2, 4):
        for c in range(7, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_tag
            cell.border = Border(
                left=border_info if c == 7 else None,
                right=border_info if c == 9 else None,
                top=border_info if r == 2 else None,
                bottom=border_info if r == 3 else None
            )

    # 2. CAIXAS DE INFORMAÇÕES (Linhas 5 a 9)
    req_number = f"#REQ-{datetime.now().strftime('%Y%m%d')}-{len(filtered_data):03d}"
    data_emissao = datetime.now().strftime("%d/%m/%Y %H:%M") + " (Horário de Brasília)"
    
    # Caixa Esquerda: Identificação da Requisição
    ws.merge_cells("B5:D5")
    ws["B5"] = "IDENTIFICAÇÃO DA REQUISIÇÃO"
    ws["B5"].font = font_sec_title
    
    ws.merge_cells("B6:D6")
    ws["B6"] = req_number
    ws["B6"].font = Font(name=font_family, size=11, bold=True, color="2A6BB8")
    
    ws["B7"] = "DATA DE EMISSÃO:"
    ws["B7"].font = font_sec_val_bold
    ws.merge_cells("C7:D7")
    ws["C7"] = data_emissao
    ws["C7"].font = font_sec_val
    
    ws["B8"] = "ORIGEM DO PEDIDO:"
    ws["B8"].font = font_sec_val_bold
    ws.merge_cells("C8:D8")
    ws["C8"] = "Laboratório Central Nova Lab"
    ws["C8"].font = font_sec_val

    
    ws.merge_cells("B9:D9")
    ws["B9"] = "Brasília - DF"
    ws["B9"].font = font_sec_val
    ws["B9"].alignment = align_left
    
    # Caixa Direita: Fornecedor Destinatário
    ws.merge_cells("F5:I5")
    ws["F5"] = "FORNECEDOR DESTINATÁRIO"
    ws["F5"].font = font_sec_title
    
    ws.merge_cells("F6:I6")
    ws["F6"] = "EssilorLuxottica Portugal / Brasil"
    ws["F6"].font = font_sec_val_bold
    
    ws["F7"] = "CNPJ / NIF:"
    ws["F7"].font = font_sec_val_bold
    ws.merge_cells("G7:I7")
    ws["G7"] = "00.000.000/0001-00"
    ws["G7"].font = font_sec_val
    
    ws["F8"] = "CANAL DE INTEGRAÇÃO:"
    ws["F8"].font = font_sec_val_bold
    ws.merge_cells("G8:I8")
    ws["G8"] = "pedidos.b2b@essilorluxottica.com.br"
    ws["G8"].font = font_sec_val

    # Desenha as bordas e fundo cinza das duas caixas
    for r in range(5, 10):
        # Caixa Esquerda
        for c in range(2, 5):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_info_box
            cell.border = Border(
                left=border_info if c == 2 else None,
                right=border_info if c == 4 else None,
                top=border_info if r == 5 else None,
                bottom=border_info if r == 9 else None
            )
        # Caixa Direita
        for c in range(6, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_info_box
            cell.border = Border(
                left=border_info if c == 6 else None,
                right=border_info if c == 9 else None,
                top=border_info if r == 5 else None,
                bottom=border_info if r == 9 else None
            )

    # 3. TÍTULO DA TABELA (Linha 11)
    ws.merge_cells("B11:I11")
    ws["B11"] = "Itens Necessários para Reposição (Abaixo do Nível de Segurança)"
    ws["B11"].font = Font(name=font_family, size=11, bold=True, color="1F4E78")
    ws["B11"].alignment = align_left
    
    # 4. CABEÇALHOS DA TABELA (Linha 13)
    headers = [
        "CÓDIGO SKU", "DESCRIÇÃO DO MODELO", "ÍNDICE", "ESFÉRICO", 
        "CILÍNDRICO", "ATUAL", "MÍN.", "QTD. SOLICITADA"
    ]
    
    for c_idx, text_header in enumerate(headers, start=2):
        cell = ws.cell(row=13, column=c_idx)
        cell.value = text_header
        cell.font = font_tbl_header
        cell.fill = fill_header
        cell.alignment = align_center if c_idx not in [2, 3] else align_left
        cell.border = cell_border
        
    # 5. DADOS DA TABELA (Linhas 14+)
    current_row = 14
    counter = 1
    
    for item in filtered_data:
        item_type = item.get('item_type', 'LENTE')
        prefix = "BLOC" if item_type == "BLOCO" else "LENS"
        sku_brand = item['brand'][:3].upper().replace(" ", "")
        sku = f"{prefix}-{int(item['refractive_index']*100)}-{sku_brand}-{counter:03d}"
        if item_type == "BLOCO":
            desc = f"[BLOCO] {item['brand']} {item['material']} {item['treatment']}"
        else:
            desc = f"[LENTE] {item['brand']} {item['material']} {item['treatment']} Ø{item['diameter']}mm"
        
        ws.cell(row=current_row, column=2, value=sku)
        ws.cell(row=current_row, column=3, value=desc)
        ws.cell(row=current_row, column=4, value=f"{item['refractive_index']:.2f}")
        ws.cell(row=current_row, column=5, value=f"{item['spherical']:+.2f}")
        ws.cell(row=current_row, column=6, value=f"{item['cylindrical']:+.2f}")
        
        # Estoque Atual com cor condicional
        stock = item['quantity_available']
        cell_stock = ws.cell(row=current_row, column=7, value=f"{stock} un")
        if stock == 0:
            cell_stock.font = font_tbl_cell_red
        elif stock <= item['reorder_point']:
            cell_stock.font = font_tbl_cell_orange
        else:
            cell_stock.font = font_tbl_cell
            
        # Mínimo (Ponto de ressuprimento)
        min_qty = int(item['reorder_point']) if item['reorder_point'] > 0 else 5
        ws.cell(row=current_row, column=8, value=f"{min_qty} un")
        
        # Quantidade Solicitada em Destaque
        suggested = item['suggested_purchase']
        cell_sugg = ws.cell(row=current_row, column=9, value=f"{suggested} un")
        cell_sugg.font = font_tbl_cell_blue
        
        is_zebra = (counter % 2 == 0)
        for c_idx in range(2, 10):
            c_cell = ws.cell(row=current_row, column=c_idx)
            if is_zebra:
                c_cell.fill = fill_zebra
            if c_idx not in [7, 9]:
                c_cell.font = font_tbl_cell_bold if c_idx in [2, 5, 6] else font_tbl_cell
            c_cell.alignment = align_center if c_idx not in [2, 3] else align_left
            c_cell.border = cell_border
            
        current_row += 1
        counter += 1
        
    # 6. DIRETRIZES DE LOGÍSTICA
    current_row += 2
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=9)
    ws.cell(row=current_row, column=2, value="Diretrizes de Logística e Transição de Estados").font = Font(name=font_family, size=11, bold=True, color="1F4E78")
    ws.cell(row=current_row, column=2).alignment = align_left
    
    # Nota automática
    current_row += 2
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row+4, end_column=9)
    note_cell = ws.cell(row=current_row, column=2)
    note_cell.value = (
        "Nota Automática do Sistema:\n"
        "1. As dioptrias listadas acima atingiram ou ultrapassaram o limite crítico de ruptura (estoque abaixo do nível de segurança). Este lote foi projetado com base na taxa de consumo diário (daily_burn_rate) para suprir os próximos 15 dias de operação.\n"
        "2. Para evitar duplicidade de pedidos por gatilhos repetidos, as respectivas células da Grade Óptica foram alteradas temporariamente para o estado STATUS_AWAITING_SUPPLIER até a recepção física da nota de faturamento."
    )
    note_cell.font = font_note_text
    note_cell.alignment = align_note
    
    border_orange_left = Side(style="medium", color="F97316")
    border_yellow_thin = Side(style="thin", color="FDE68A")
    
    for r in range(current_row, current_row + 5):
        for c in range(2, 10):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_note
            cell.border = Border(
                left=border_orange_left if c == 2 else None,
                right=border_yellow_thin if c == 9 else None,
                top=border_yellow_thin if r == current_row else None,
                bottom=border_yellow_thin if r == current_row + 4 else None
            )
            
    # Ajusta larguras e alturas
    ws.column_dimensions['A'].width = 3
    col_widths = {
        'B': 18, 'C': 35, 'D': 10, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 16
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[13].height = 25
    for r in range(14, current_row - 4):
        ws.row_dimensions[r].height = 20
        
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
