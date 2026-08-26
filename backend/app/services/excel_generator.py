import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate_billing_excel(cycle) -> bytes:
    """
    Gera uma planilha Excel formatada profissionalmente contendo
    o detalhamento de OSs faturadas em um lote de fechamento, discriminando
    os produtos, serviços técnicos e tratamentos conforme o catálogo (Anexo 1 & Anexo 2).
    """
    # Coletamos os dados das OSs e seus itens detalhados
    data = []
    for item in cycle.items:
        os_num = item.os_number or "N/A"
        client = item.client_name or "Consumidor Final"
        created_str = item.created_at.strftime("%d/%m/%Y %H:%M:%S") if hasattr(item, "created_at") and item.created_at else "-"
        detailed_list = getattr(item, "detailed_items", []) or []

        if detailed_list:
            for detail in detailed_list:
                d_name = getattr(detail, "name", "") or "Item"
                d_desc = getattr(detail, "description", "") or "-"
                d_type = getattr(detail, "item_type", "") or "Serviço"
                d_qty = getattr(detail, "quantity", 1)
                d_uprice = float(getattr(detail, "unit_price", 0.0) or 0.0)
                d_tprice = float(getattr(detail, "total_price", 0.0) or (d_uprice * d_qty))

                data.append({
                    "Código da OS": os_num,
                    "Paciente / Cliente Final": client,
                    "Item / Serviço": d_name,
                    "Descrição do Catálogo": d_desc,
                    "Tipo": d_type,
                    "Quantidade": f"{d_qty} un" if d_type == "Lente" else str(d_qty),
                    "Valor Unitário (R$)": d_uprice,
                    "Valor Total (R$)": d_tprice,
                    "Data do Fechamento": created_str
                })
        else:
            # Fallback para OSs legadas
            lens_txt = getattr(item, "lens_type", "Lente Padrão Laboratorial")
            lens_price_val = float(getattr(item, "lens_price", None) if getattr(item, "lens_price", None) is not None else (item.amount or 0.0))
            data.append({
                "Código da OS": os_num,
                "Paciente / Cliente Final": client,
                "Item / Serviço": lens_txt,
                "Descrição do Catálogo": "Lente Oftálmica Visão Simples / Digital",
                "Tipo": "Lente",
                "Quantidade": "2 un",
                "Valor Unitário (R$)": round(lens_price_val / 2, 2) if lens_price_val > 0 else 0.0,
                "Valor Total (R$)": lens_price_val,
                "Data do Fechamento": created_str
            })
            
            services_txt = getattr(item, "services", None)
            service_price_val = float(getattr(item, "service_price", 0.0) or 0.0)
            if services_txt:
                data.append({
                    "Código da OS": os_num,
                    "Paciente / Cliente Final": client,
                    "Item / Serviço": services_txt,
                    "Descrição do Catálogo": "Montagem e Acabamento de Precisão",
                    "Tipo": "Serviço",
                    "Quantidade": "1",
                    "Valor Unitário (R$)": service_price_val,
                    "Valor Total (R$)": service_price_val,
                    "Data do Fechamento": created_str
                })

            treatments_txt = getattr(item, "treatments", None)
            treatment_price_val = float(getattr(item, "treatment_price", 0.0) or 0.0)
            if treatments_txt and treatments_txt != "Incolor / Sem Tratamento":
                data.append({
                    "Código da OS": os_num,
                    "Paciente / Cliente Final": client,
                    "Item / Serviço": treatments_txt,
                    "Descrição do Catálogo": "Tratamento de Superfície Lente",
                    "Tipo": "Tratamento",
                    "Quantidade": "1",
                    "Valor Unitário (R$)": treatment_price_val,
                    "Valor Total (R$)": treatment_price_val,
                    "Data do Fechamento": created_str
                })

    df = pd.DataFrame(data)
    
    # Criamos o buffer em memória
    output = io.BytesIO()
    
    # Escrevemos com pandas
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Faturamento Detalhado")
        
        # Obter o workbook e a aba gerada para estilização
        workbook = writer.book
        worksheet = writer.sheets["Faturamento Detalhado"]
        
        # Ajusta a largura das colunas de forma inteligente
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        # Estilos premium para o Excel
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul clássico
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_font = Font(name="Calibri", size=10)
        align_left = Alignment(horizontal="left", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_side = Side(border_style="thin", color="D1D5DB")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Aplica estilo no cabeçalho
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border_all
            
        # Aplica estilo e formatação decimal/moeda nas células de dados
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = border_all
                
                # Alinhamentos e formatações específicos por coluna
                if col_idx in [1, 5, 6, 9]: # OS, Tipo, Qtd e Data
                    cell.alignment = align_center
                elif col_idx in [2, 3, 4]: # Paciente, Item/Serviço, Descrição
                    cell.alignment = align_left
                elif col_idx in [7, 8]: # Valores (Unitário e Total)
                    cell.alignment = align_right
                    cell.number_format = '"R$"#,##0.00'
                    
        # Linha de total no final da planilha
        total_row_idx = len(df) + 3
        worksheet.cell(row=total_row_idx, column=1, value="TOTAL GERAL FATURADO").font = Font(name="Calibri", size=11, bold=True)
        worksheet.cell(row=total_row_idx, column=1).alignment = align_left
        
        total_cell = worksheet.cell(row=total_row_idx, column=8, value=float(cycle.total_amount))
        total_cell.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
        total_cell.alignment = align_right
        total_cell.number_format = '"R$"#,##0.00'
        
        # Borda dupla inferior e simples superior para a linha de total
        top_side = Side(border_style="thin", color="1F4E78")
        double_bottom = Side(border_style="double", color="1F4E78")
        total_border = Border(top=top_side, bottom=double_bottom)
        
        for col_idx in range(1, len(df.columns) + 1):
            worksheet.cell(row=total_row_idx, column=col_idx).border = total_border
            
    return output.getvalue()
